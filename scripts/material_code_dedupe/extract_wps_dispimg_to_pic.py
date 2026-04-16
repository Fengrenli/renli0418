#!/usr/bin/env python3
"""
从 WPS 表格中按 =DISPIMG("ID_…",1) 抠图并落到 PIC/<pic_folder>/。

与「仅解压 xl/media + 按序号对齐」不同：本脚本解析
  - xl/cellimages.xml（或同类节点）里 cNvPr@name = DISPIMG 里的 ID
  - a:blip 的 r:embed → xl/_rels/cellimages.xml.rels → xl/media/…
从而 **ID 与二进制一一对应**，再按表内 **code**（及可选 name 后缀）命名。

依赖：openpyxl（读公式）、标准库 zipfile/xml。

示例（吼堂前厅杂件，仅输出数据库用纯码文件名）：
  python extract_wps_dispimg_to_pic.py \\
    --workbook "C:/Users/Admin/Desktop/material/new/houtang/material-excel/前厅杂件.bak_before_code.xlsx" \\
    --pic-dir "C:/Users/Admin/Desktop/material/new/houtang/PIC/前厅杂件" \\
    --naming code

若需与其它 AI 相同「码_名称」可读文件名：
  python extract_wps_dispimg_to_pic.py ... --naming code_and_name
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
import pandas as pd

DISPIMG_RE = re.compile(
    r'DISPIMG\s*\(\s*["\'](ID_[A-F0-9]+)["\']',
    re.I,
)
INVALID_FS = re.compile(r'[<>:"/\\|?*\n\r\t]')


def norm_header(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def find_col(ws: openpyxl.worksheet.worksheet.Worksheet, want: str) -> int | None:
    for c in range(1, ws.max_column + 1):
        if norm_header(ws.cell(1, c).value) == want:
            return c
    return None


def find_col_any(ws: openpyxl.worksheet.worksheet.Worksheet, wants: tuple[str, ...]) -> int | None:
    want_set = {w.strip().lower() for w in wants if w and w.strip()}
    for c in range(1, ws.max_column + 1):
        h = norm_header(ws.cell(1, c).value)
        if h in want_set:
            return c
    return None


def local_tag(tag: str) -> str:
    if "}" in tag:
        return tag.rsplit("}", 1)[-1]
    return tag


def get_attrib_embed(elem: ET.Element) -> str | None:
    for k, v in elem.attrib.items():
        if k.endswith("}embed") or k == "embed":
            return v
    return None


def parse_cellimages_id_to_embed(zipf: zipfile.ZipFile) -> dict[str, str]:
    """DISPIMG 公式中的 ID_xxx -> relationship Id (rId...)"""
    names_lower = {n.lower(): n for n in zipf.namelist()}
    ci_key = next(
        (names_lower[k] for k in ("xl/cellimages.xml", "xl/cellImages.xml") if k in names_lower),
        None,
    )
    if not ci_key:
        return {}
    root = ET.fromstring(zipf.read(ci_key))
    out: dict[str, str] = {}
    for node in root.iter():
        if local_tag(node.tag) != "cellImage":
            continue
        cnv_name: str | None = None
        embed_rid: str | None = None
        for d in node.iter():
            lt = local_tag(d.tag)
            if lt == "cNvPr":
                nm = d.get("name")
                if nm and nm.startswith("ID_"):
                    cnv_name = nm
            if lt == "blip":
                e = get_attrib_embed(d)
                if e:
                    embed_rid = e
        if cnv_name and embed_rid:
            out[cnv_name] = embed_rid
    return out


def parse_cellimages_order(zipf: zipfile.ZipFile) -> list[str]:
    """按 cellimages.xml 出现顺序返回 ID_ 列表（用于无 DISPIMG 公式时的顺序兜底）。"""
    names_lower = {n.lower(): n for n in zipf.namelist()}
    ci_key = next(
        (names_lower[k] for k in ("xl/cellimages.xml", "xl/cellImages.xml") if k in names_lower),
        None,
    )
    if not ci_key:
        return []
    root = ET.fromstring(zipf.read(ci_key))
    out: list[str] = []
    for node in root.iter():
        if local_tag(node.tag) != "cellImage":
            continue
        for d in node.iter():
            if local_tag(d.tag) == "cNvPr":
                nm = d.get("name")
                if nm and nm.startswith("ID_"):
                    out.append(nm)
                    break
    return out


def parse_cellimages_id_to_anchor(zipf: zipfile.ZipFile) -> dict[str, tuple[int, int]]:
    """
    尝试从 cellimages.xml 解析每个 cellImage 的锚点行列（1-based）。
    不同 WPS/版本结构略有差异：这里做 best-effort，找含 row/col 的 from 节点。
    """
    names_lower = {n.lower(): n for n in zipf.namelist()}
    ci_key = next(
        (names_lower[k] for k in ("xl/cellimages.xml", "xl/cellImages.xml") if k in names_lower),
        None,
    )
    if not ci_key:
        return {}
    root = ET.fromstring(zipf.read(ci_key))
    out: dict[str, tuple[int, int]] = {}

    def find_from_row_col(node: ET.Element) -> tuple[int, int] | None:
        r_val: int | None = None
        c_val: int | None = None
        for d in node.iter():
            lt = local_tag(d.tag)
            if lt == "row" and d.text and d.text.strip().isdigit():
                r_val = int(d.text.strip())
            elif lt == "col" and d.text and d.text.strip().isdigit():
                c_val = int(d.text.strip())
        if r_val is None or c_val is None:
            return None
        # OOXML anchor 通常是 0-based，这里转成 Excel 1-based 行列
        return (r_val + 1, c_val + 1)

    for node in root.iter():
        if local_tag(node.tag) != "cellImage":
            continue
        cnv_name: str | None = None
        for d in node.iter():
            if local_tag(d.tag) == "cNvPr":
                nm = d.get("name")
                if nm and nm.startswith("ID_"):
                    cnv_name = nm
                    break
        if not cnv_name:
            continue
        rc = find_from_row_col(node)
        if rc:
            out[cnv_name] = rc
    return out


def parse_cellimages_rels(zipf: zipfile.ZipFile) -> dict[str, str]:
    """rId -> xl/media/xxx 相对路径（正斜杠）"""
    rel_key = "xl/_rels/cellimages.xml.rels"
    names_lower = {n.lower(): n for n in zipf.namelist()}
    rk = names_lower.get(rel_key.lower())
    if not rk:
        return {}
    root = ET.fromstring(zipf.read(rk))
    out: dict[str, str] = {}
    for rel in root.iter():
        if local_tag(rel.tag) != "Relationship":
            continue
        rid = rel.get("Id")
        tgt = rel.get("Target")
        if rid and tgt:
            out[rid] = tgt.replace("\\", "/")
    return out


def resolve_media_path(target: str) -> str:
    """cellimages rel Target 如 ../media/image1.png -> xl/media/image1.png"""
    t = target.strip().replace("\\", "/")
    if t.startswith("../"):
        t = "xl/" + t[3:]
    elif not t.startswith("xl/"):
        t = "xl/" + t.lstrip("/")
    return t


def build_dispimg_id_to_bytes(zipf: zipfile.ZipFile) -> dict[str, tuple[str, bytes]]:
    """ID_xxx -> (member_name_in_zip, file_bytes)"""
    id2rid = parse_cellimages_id_to_embed(zipf)
    rid2tgt = parse_cellimages_rels(zipf)
    out: dict[str, tuple[str, bytes]] = {}
    for img_id, rid in id2rid.items():
        tgt = rid2tgt.get(rid)
        if not tgt:
            continue
        member = resolve_media_path(tgt)
        names_lower = {n.lower(): n for n in zipf.namelist()}
        key = member.lower()
        if key not in names_lower:
            continue
        real = names_lower[key]
        out[img_id] = (real, zipf.read(real))
    return out


def safe_name_fragment(name: str, max_len: int = 80) -> str:
    t = INVALID_FS.sub("_", (name or "").strip())
    if len(t) > max_len:
        t = t[:max_len]
    return t


def main() -> None:
    ap = argparse.ArgumentParser(description="WPS DISPIMG → PIC 按 code 命名")
    ap.add_argument("--workbook", "-w", type=Path, required=True, help="xlsx 路径")
    ap.add_argument(
        "--pic-dir",
        "-o",
        type=Path,
        required=True,
        help="输出目录，如 .../PIC/前厅杂件",
    )
    ap.add_argument(
        "--naming",
        choices=("code", "code_and_name"),
        default="code",
        help="code=仅 HW10-xxx.ext（与 OSS/库一致）；code_and_name=HW10-xxx_名称.ext",
    )
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--mapping-json",
        type=Path,
        default=None,
        help="可选，写入 image_mapping.json（DISPIMG 行信息）",
    )
    ap.add_argument(
        "--mapping-csv",
        type=Path,
        default=None,
        help="可选：materials_cleaned_*.csv；当表内 code 为空时按 pic_folder 的 code 顺序兜底命名",
    )
    ap.add_argument(
        "--pic-folder",
        type=str,
        default="",
        help="配合 --mapping-csv 指定 pic_folder（默认取输出目录名）",
    )
    args = ap.parse_args()

    wb_path = args.workbook.resolve()
    if not wb_path.is_file():
        sys.exit(f"文件不存在: {wb_path}")

    pic_dir = args.pic_dir.resolve()

    with zipfile.ZipFile(wb_path, "r") as zf:
        id_to_blob = build_dispimg_id_to_bytes(zf)
        id_to_anchor = parse_cellimages_id_to_anchor(zf)
        id_order = parse_cellimages_order(zf)

    if not id_to_blob:
        print(
            "[ERR] 未能从 cellimages.xml / cellimages.xml.rels 建立 DISPIMG ID 映射。"
            " 可能不是 WPS 单元格图、或该本为 Excel 另存导致无 cellimages。"
            " 用 rebuild_houtang_qianting.ps1 时会自动换主/备份 xlsx 并重试，仍失败则回退 openpyxl 嵌入图。",
            file=sys.stderr,
        )
        sys.exit(2)

    print(f"[INFO] cellimages 映射: {len(id_to_blob)} 个 ID → media", file=sys.stderr)

    wb = openpyxl.load_workbook(wb_path, data_only=False, rich_text=False)
    ws = wb.active
    # 兼容中英表头
    img_col = find_col_any(ws, ("image", "图片", "图片url", "图片 url", "image url", "image_url"))
    code_col = find_col_any(ws, ("code", "编码", "物料编码", "ITEM CODE", "item code"))
    name_col = find_col_any(ws, ("name", "名称", "物料名称", "产品名称", "product name"))
    if img_col is None or code_col is None:
        sys.exit("表头需要 image/图片 与 code/编码 列")

    rows_out: list[dict[str, object]] = []
    missing_id: list[tuple[int, str, str]] = []
    written = 0
    skipped_dup = 0

    # 可选：从 mapping csv 读取该 pic_folder 的 code 顺序，作为 code 为空时兜底
    fallback_codes: list[str] = []
    fallback_i = 0
    if args.mapping_csv is not None and args.mapping_csv.is_file():
        try:
            mc = pd.read_csv(args.mapping_csv, dtype=str, keep_default_na=False)
            mc.columns = [str(c).strip().lower() for c in mc.columns]
            if "pic_folder" in mc.columns and "code" in mc.columns:
                pf = args.pic_folder.strip() if args.pic_folder.strip() else pic_dir.name
                sub = mc[mc["pic_folder"].astype(str).str.strip() == pf].copy()
                if "id" in sub.columns:
                    sub["_idn"] = pd.to_numeric(sub["id"], errors="coerce").fillna(0).astype(int)
                    sub = sub.sort_values("_idn", kind="mergesort")
                fallback_codes = [str(x).strip() for x in sub["code"].tolist() if str(x).strip()]
                if fallback_codes:
                    print(f"[INFO] code 兜底队列: pic_folder={pf!r}, n={len(fallback_codes)}", file=sys.stderr)
        except Exception as e:
            print(f"[INFO] 读取 mapping-csv 失败，忽略 code 兜底: {e}", file=sys.stderr)

    for r in range(2, ws.max_row + 1):
        raw_img = ws.cell(r, img_col).value
        if raw_img is None or str(raw_img).strip() == "":
            continue
        m = DISPIMG_RE.search(str(raw_img))
        if not m:
            continue
        img_id = m.group(1)
        code_cell = ws.cell(r, code_col).value
        code = "" if code_cell is None else str(code_cell).strip()
        if not code and fallback_i < len(fallback_codes):
            code = fallback_codes[fallback_i]
            fallback_i += 1
        if not code:
            continue
        nm = ""
        if name_col:
            nv = ws.cell(r, name_col).value
            nm = "" if nv is None else str(nv).strip()

        blob_info = id_to_blob.get(img_id)
        row_rec: dict[str, object] = {
            "row": r,
            "dispimg_id": img_id,
            "code": code,
            "name": nm,
        }
        if not blob_info:
            missing_id.append((r, code, img_id))
            row_rec["status"] = "id_not_in_cellimages"
            rows_out.append(row_rec)
            continue

        _member, data = blob_info
        ext = Path(_member).suffix.lower()
        if ext not in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
            ext = ".png"

        if args.naming == "code_and_name" and nm:
            base = f"{code}_{safe_name_fragment(nm)}"
        else:
            base = code

        out_name = f"{base}{ext}"
        out_path = pic_dir / out_name
        row_rec["out_file"] = out_name
        row_rec["status"] = "ok"

        if args.dry_run:
            rows_out.append(row_rec)
            written += 1
            continue

        pic_dir.mkdir(parents=True, exist_ok=True)
        if out_path.is_file():
            skipped_dup += 1
            row_rec["status"] = "skip_existing"
            rows_out.append(row_rec)
            continue

        out_path.write_bytes(data)
        written += 1
        rows_out.append(row_rec)
        print(f"  {out_name}  <- 行{r}  {img_id}")

    # 若表内没有 DISPIMG 公式（常见：WPS 单元格图，但未在 image 列落公式），尝试用 cellimages.xml 的锚点行直接取 code
    if written == 0 and id_to_anchor:
        wrote2 = 0
        for img_id, (r1, _c1) in sorted(id_to_anchor.items(), key=lambda x: x[1]):
            blob_info = id_to_blob.get(img_id)
            if not blob_info:
                continue
            code_cell = ws.cell(r1, code_col).value
            code = "" if code_cell is None else str(code_cell).strip()
            if not code and fallback_i < len(fallback_codes):
                code = fallback_codes[fallback_i]
                fallback_i += 1
            if not code:
                continue
            nm = ""
            if name_col:
                nv = ws.cell(r1, name_col).value
                nm = "" if nv is None else str(nv).strip()

            _member, data = blob_info
            ext = Path(_member).suffix.lower()
            if ext not in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
                ext = ".png"

            if args.naming == "code_and_name" and nm:
                base = f"{code}_{safe_name_fragment(nm)}"
            else:
                base = code
            out_name = f"{base}{ext}"
            out_path = pic_dir / out_name
            if not args.dry_run:
                pic_dir.mkdir(parents=True, exist_ok=True)
                if out_path.is_file():
                    continue
                out_path.write_bytes(data)
            wrote2 += 1
        if wrote2 > 0:
            written = wrote2
            print(f"[INFO] fallback: 从 cellimages 锚点行导出 {wrote2} 张（未依赖 DISPIMG 公式）", file=sys.stderr)

    # 顺序兜底：若既无 DISPIMG 公式也无可解析锚点（WPS cellimages 仅有坐标 off/ext），则只能按顺序对齐。
    if written == 0 and id_order:
        codes_in_sheet: list[tuple[int, str, str]] = []
        for r in range(2, ws.max_row + 1):
            code_cell = ws.cell(r, code_col).value
            code = "" if code_cell is None else str(code_cell).strip()
            if not code and fallback_i < len(fallback_codes):
                code = fallback_codes[fallback_i]
                fallback_i += 1
            if not code:
                continue
            nm = ""
            if name_col:
                nv = ws.cell(r, name_col).value
                nm = "" if nv is None else str(nv).strip()
            codes_in_sheet.append((r, code, nm))

        pairs = list(zip(id_order, codes_in_sheet))
        if pairs:
            print(
                f"[WARN] 未找到 DISPIMG 公式且无法解析锚点行；将按 cellimages 出现顺序对齐到 code 行顺序（可能需人工抽查）。",
                file=sys.stderr,
            )
        wrote3 = 0
        for img_id, (r, code, nm) in pairs:
            blob_info = id_to_blob.get(img_id)
            if not blob_info:
                continue
            _member, data = blob_info
            ext = Path(_member).suffix.lower()
            if ext not in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
                ext = ".png"
            if args.naming == "code_and_name" and nm:
                base = f"{code}_{safe_name_fragment(nm)}"
            else:
                base = code
            out_name = f"{base}{ext}"
            out_path = pic_dir / out_name
            if not args.dry_run:
                pic_dir.mkdir(parents=True, exist_ok=True)
                if out_path.is_file():
                    continue
                out_path.write_bytes(data)
            wrote3 += 1
            rows_out.append({"row": r, "dispimg_id": img_id, "code": code, "name": nm, "out_file": out_name, "status": "fallback_order"})
        if wrote3 > 0:
            written = wrote3
            print(f"[INFO] fallback: 顺序对齐导出 {wrote3} 张", file=sys.stderr)

    if missing_id:
        print(f"[WARN] {len(missing_id)} 行 DISPIMG ID 在 cellimages 中无对应 media", file=sys.stderr)
        for r, c, i in missing_id[:15]:
            print(f"       行{r} code={c!r} id={i}", file=sys.stderr)
        if len(missing_id) > 15:
            print(f"       ... 共 {len(missing_id)} 行", file=sys.stderr)

    map_path = args.mapping_json
    if map_path is None:
        map_path = pic_dir / "image_mapping.json"
    if not args.dry_run and map_path:
        map_path.parent.mkdir(parents=True, exist_ok=True)
        with open(map_path, "w", encoding="utf-8") as f:
            json.dump(rows_out, f, ensure_ascii=False, indent=2)
        print(f"[INFO] 映射已写: {map_path.resolve()}", file=sys.stderr)

    print(
        f"完成: 写出 {written}，跳过已存在 {skipped_dup}，"
        f"DISPIMG 行共 {len(rows_out)}，cellimages 库 {len(id_to_blob)} 个 ID",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
