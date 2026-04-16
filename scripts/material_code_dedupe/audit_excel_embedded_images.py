#!/usr/bin/env python3
"""
审计 .xlsx 嵌入图（openpyxl），回答「为什么 media 里 150+ 张图 ≠ 148 行物料」。

要点：
  - 解压所见的 xl/media/*.png 是**资源池**，文件个数不能当作「每行一张图」。
  - 本脚本按**锚点 (行, 列)** 列出 openpyxl 能读到的嵌入对象，并做 **SHA256** 分组（完全相同的文件只算一种字节）。

输出 CSV 含：工作簿、图序号、锚点行列、是否落在表头为 image 的列、sha256、字节数、
以及「与上一张图是否同哈希」（便于肉眼看重复链）。

可选：传入 materials_cleaned*.csv，在 stderr 打印该 pic_folder 下有码行数 vs 嵌入图数对照。

用法：
  python audit_excel_embedded_images.py \\
    --workbook "C:/Users/Admin/Desktop/material/new/houtang/material-excel/前厅杂件.bak_before_code.xlsx" \\
    -o "E:/renli0418/excel_embed_audit_前厅杂件.csv"

  python audit_excel_embedded_images.py -w "...xlsx" --mapping "E:/renli0418/materials_cleaned_houtang.csv"
"""
from __future__ import annotations

import argparse
import hashlib
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd

INVALID_FS = re.compile(r'[<>:"/\\|?*]')


def norm_header(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def header_col(ws: openpyxl.worksheet.worksheet.Worksheet, want: str) -> int | None:
    for c in range(1, ws.max_column + 1):
        if norm_header(ws.cell(1, c).value) == want:
            return c
    return None


def image_sort_key(img: object) -> tuple[int, int]:
    a = img.anchor
    fr = a._from
    return (int(fr.row), int(fr.col))


def main() -> None:
    ap = argparse.ArgumentParser(description="审计 xlsx 嵌入图锚点与重复（勿用 xl/media 个数当行数）")
    ap.add_argument(
        "-w",
        "--workbook",
        type=Path,
        required=True,
        help="单个 .xlsx 路径",
    )
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="输出 CSV（默认：工作簿同目录下 excel_embed_audit_<stem>.csv）",
    )
    ap.add_argument(
        "--mapping",
        type=Path,
        default=None,
        help="可选 materials_cleaned*.csv，用于对照 pic_folder 有码行数",
    )
    ap.add_argument("--sheet", type=int, default=0, help="0-based 工作表索引，默认 0")
    args = ap.parse_args()

    path = args.workbook.resolve()
    if not path.is_file():
        sys.exit(f"文件不存在: {path}")
    if path.suffix.lower() != ".xlsx":
        sys.exit("仅支持 .xlsx（openpyxl）")

    stem_key = path.stem.strip().rstrip(".")
    if stem_key.endswith(".bak_before_code"):
        stem_key = stem_key[: -len(".bak_before_code")]

    out = args.output
    if out is None:
        out = path.parent / f"excel_embed_audit_{stem_key}.csv"

    wb = openpyxl.load_workbook(path, data_only=True, rich_text=False)
    if args.sheet >= len(wb.sheetnames):
        sys.exit(f"sheet 索引越界: {args.sheet} >= {len(wb.sheetnames)}")
    ws = wb[wb.sheetnames[args.sheet]]

    image_col = header_col(ws, "image")
    name_col = header_col(ws, "name")
    code_col = header_col(ws, "code")

    images = sorted(ws._images, key=image_sort_key)
    rows_out: list[dict[str, object]] = []
    prev_hash = ""
    hash_to_indices: dict[str, list[int]] = defaultdict(list)

    for i, img in enumerate(images):
        fr = img.anchor._from
        row_1 = int(fr.row) + 1
        col_1 = int(fr.col) + 1
        on_image_col = image_col is not None and col_1 == image_col
        try:
            data = img._data()
        except Exception as e:
            rows_out.append(
                {
                    "workbook": path.name,
                    "img_index": i + 1,
                    "anchor_row": row_1,
                    "anchor_col": col_1,
                    "on_image_column": on_image_col,
                    "sha256": "",
                    "bytes": 0,
                    "same_as_prev": "",
                    "error": str(e),
                }
            )
            continue
        h = hashlib.sha256(data).hexdigest()
        hash_to_indices[h].append(i + 1)
        same_prev = "yes" if h == prev_hash and prev_hash else ""
        prev_hash = h
        rows_out.append(
            {
                "workbook": path.name,
                "img_index": i + 1,
                "anchor_row": row_1,
                "anchor_col": col_1,
                "on_image_column": on_image_col,
                "sha256": h,
                "bytes": len(data),
                "same_as_prev": same_prev,
                "error": "",
            }
        )

    # 按锚点行统计（多张图落在同一行）
    by_row: dict[int, list[int]] = defaultdict(list)
    for i, img in enumerate(images):
        fr = img.anchor._from
        by_row[int(fr.row) + 1].append(i + 1)

    multi_row = {r: idxs for r, idxs in by_row.items() if len(idxs) > 1}
    dup_groups = {h: idxs for h, idxs in hash_to_indices.items() if len(idxs) > 1}

    pd.DataFrame(rows_out).to_csv(out, index=False, encoding="utf-8-sig")
    print(f"已写入: {out.resolve()}")
    print(f"嵌入对象数（openpyxl）: {len(images)}")
    print(f"唯一 SHA256 数: {len(hash_to_indices)}")
    if dup_groups:
        print(f"完全重复（同字节）的图组数: {len(dup_groups)}（见 CSV 按 sha256 筛选）")
    if multi_row:
        print(
            f"同一数据行锚点多图: {len(multi_row)} 行（行号: {sorted(multi_row.keys())[:12]}"
            f"{'...' if len(multi_row) > 12 else ''}）",
            file=sys.stderr,
        )

    if args.mapping and args.mapping.is_file():
        mc = pd.read_csv(args.mapping, dtype=str, keep_default_na=False)
        mc.columns = [str(c).strip().lower() for c in mc.columns]
        if "pic_folder" in mc.columns and "code" in mc.columns:
            sub = mc[mc["pic_folder"].astype(str).str.strip() == stem_key]
            c = sub["code"].astype(str).str.strip()
            n_codes = int((c.ne("") & ~c.str.lower().isin(("nan", "none"))).sum())
            print(
                f"对照 mapping: pic_folder={stem_key!r} 有码行数={n_codes}，"
                f"嵌入图数={len(images)}（多出的通常来自重复嵌入或非 image 列/非数据行）",
                file=sys.stderr,
            )
            if n_codes and len(images) != n_codes:
                print(
                    "  建议：以「锚点排序后前 n_codes 张」或 extract_images_from_material_excel.py 的告警为准；"
                    "不要用 xl/media 文件个数对齐行号。",
                    file=sys.stderr,
                )

    if name_col:
        n_data = 0
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, name_col).value
            if v is not None and str(v).strip():
                n_data += 1
        print(f"表内非空 name 行数（约）: {n_data}", file=sys.stderr)


if __name__ == "__main__":
    main()
