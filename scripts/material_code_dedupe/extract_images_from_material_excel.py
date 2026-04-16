#!/usr/bin/env python3
"""
从 material-excel/*.xlsx 中读取**嵌入图片**，按 materials_cleaned.csv 写入 PIC/<pic_folder>/<新码>.png。

前提：
  - 已跑过 merge + clean，得到 materials_cleaned_*.csv（含 id、code、code_before_cleanup、pic_folder）。
  - 每个分表一个 xlsx，文件名（不含扩展名）= pic_folder；图片自动落到 PIC/该文件夹/。
  - **仅处理有「新 code」的行**：空 code 不占位、不配图；图按嵌入对象 (行,列) 排序后与这些行**按顺序 zip**。
  - position 模式：表内可无「code」列（仅靠合并 CSV 行序对齐）；cell 模式需要表头 code 列做旧码匹配。
  - **老版 .xls 无法用 openpyxl 读嵌入图**，请另存为 .xlsx 后再跑本脚本。
  - **WPS =DISPIMG(...)** 占格的图**不是** OOXML 嵌入对象，openpyxl 的 ws._images 里**没有**，本脚本导不出；需改用真嵌入图或本地 PIC 按码命名。

用法：
  python extract_images_from_material_excel.py \\
    --excel-dir "C:/Users/Admin/Desktop/material/new/houtang/material-excel" \\
    --mapping "E:/renli0418/materials_cleaned_houtang.csv" \\
    --pic-root "C:/Users/Admin/Desktop/material/new/houtang/PIC"

  python extract_images_from_material_excel.py ... --dry-run
  python extract_images_from_material_excel.py ... --match cell

  # 只重做「前厅杂件」且按单元格 code 对齐（避免 image 顺序错位）：
  python extract_images_from_material_excel.py ... --match cell --only-stem 前厅杂件 --prefer-bak
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict, deque
from pathlib import Path

import openpyxl
import pandas as pd

INVALID_FS = re.compile(r'[<>:"/\\|?*]')


def safe_stem(code: str) -> str:
    t = (code or "").strip()
    t = INVALID_FS.sub("_", t)
    return t or "unnamed"


def norm_header(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def find_code_col(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    for c in range(1, ws.max_column + 1):
        if norm_header(ws.cell(1, c).value) == "code":
            return c
    return None


def image_sort_key(img: object) -> tuple[int, int]:
    a = img.anchor
    fr = a._from
    return (int(fr.row), int(fr.col))


def count_dispimg_formula_cells(path: Path) -> int:
    """统计含 WPS DISPIMG 公式的单元格数（需 data_only=False）。"""
    wb = openpyxl.load_workbook(path, data_only=False, rich_text=False)
    try:
        ws = wb.active
        n = 0
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "DISPIMG" in v.upper():
                    n += 1
    finally:
        wb.close()
    return n


def row_has_new_code(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    return s.ne("") & ~s.str.lower().isin(("nan", "none"))


def pic_folder_stem_from_path(path: Path) -> str:
    """文件名 stem → 与 CSV pic_folder / PIC 子目录一致（去掉 .bak_before_code 后缀）。"""
    s = path.stem.strip().rstrip(".")
    suf = ".bak_before_code"
    if len(s) > len(suf) and s.lower().endswith(suf.lower()):
        return s[: -len(suf)].rstrip(".")
    return s


def dedupe_workbooks_one_per_folder(files: list[Path], prefer_bak: bool) -> list[Path]:
    """同 pic_folder 同时存在 前厅杂件.xlsx 与 前厅杂件.bak_before_code.xlsx 时只保留一个，避免重复导出。"""
    g: dict[str, list[Path]] = defaultdict(list)
    for p in files:
        g[pic_folder_stem_from_path(p)].append(p)
    out: list[Path] = []
    for key in sorted(g.keys()):
        grp = sorted(g[key], key=lambda x: x.name.lower())
        if len(grp) == 1:
            out.append(grp[0])
            continue
        baks = [p for p in grp if ".bak_before_code" in p.name.lower()]
        mains = [p for p in grp if ".bak_before_code" not in p.name.lower()]
        if prefer_bak:
            chosen = baks[0] if baks else grp[0]
        else:
            chosen = mains[0] if mains else grp[0]
        print(
            f"[INFO] pic_folder={key!r} 有多个工作簿，选用: {chosen.name}",
            file=sys.stderr,
        )
        out.append(chosen)
    return sorted(out, key=lambda p: p.name.lower())


def main() -> None:
    ap = argparse.ArgumentParser(description="从 material-excel xlsx 导出嵌入图到 PIC 子文件夹")
    ap.add_argument(
        "--excel-dir",
        type=Path,
        default=Path(r"C:\Users\Admin\Desktop\material\new\material-excel"),
    )
    ap.add_argument(
        "--mapping",
        type=Path,
        default=Path(__file__).resolve().parents[2] / "materials_cleaned.csv",
        help="materials_cleaned.csv（需含 id, code, code_before_cleanup, pic_folder）",
    )
    ap.add_argument(
        "--pic-root",
        type=Path,
        default=Path(r"C:\Users\Admin\Desktop\material\new\PIC"),
    )
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--skip-existing", action="store_true", help="目标文件已存在则跳过")
    ap.add_argument(
        "--match",
        choices=("position", "cell", "hybrid"),
        default="position",
        help="position: 图与映射行按顺序 zip；cell: 按锚点行 code 匹配 code_before_cleanup；"
        "hybrid: 优先按 excel_src_row（merge 写入）与图锚点行对齐，再 cell，"
        "再锚点行 code 为空时按 id 顺序兜底（需重跑 merge 后才有 excel_src_row）",
    )
    ap.add_argument(
        "--only-stem",
        default="",
        help="只处理该 pic_folder（如 前厅杂件），与主文件名 stem 一致，不含 .bak_before_code",
    )
    ap.add_argument(
        "--prefer-bak",
        action="store_true",
        help="同目录既有主 xlsx 又有 *.bak_before_code.xlsx 时优先读备份（与 merge --use-bak-if-present 一致）",
    )
    args = ap.parse_args()

    if not args.excel_dir.is_dir():
        sys.exit(f"目录不存在: {args.excel_dir}")
    if not args.mapping.is_file():
        sys.exit(f"找不到: {args.mapping}")

    mc = pd.read_csv(args.mapping, dtype=str, keep_default_na=False)
    mc.columns = [str(c).strip().lower() for c in mc.columns]
    for col in ("id", "code", "pic_folder"):
        if col not in mc.columns:
            sys.exit(f"{args.mapping} 缺少列: {col}")
    if "code_before_cleanup" not in mc.columns:
        sys.exit(f"{args.mapping} 缺少 code_before_cleanup（请用当前流程重新生成 materials_cleaned.csv）")

    xlsx_files = sorted(
        (p for p in args.excel_dir.glob("*.xlsx") if not p.name.startswith("~$")),
        key=lambda p: p.name,
    )
    if not xlsx_files:
        sys.exit(
            f"未找到 .xlsx: {args.excel_dir}（.xls 不支持抠嵌入图，请在 Excel 中另存为 .xlsx）"
        )

    xlsx_files = dedupe_workbooks_one_per_folder(xlsx_files, args.prefer_bak)
    if args.only_stem.strip():
        key = args.only_stem.strip()
        xlsx_files = [p for p in xlsx_files if pic_folder_stem_from_path(p) == key]
        if not xlsx_files:
            sys.exit(f"--only-stem={key!r} 无匹配工作簿（已按 pic_folder 去重后筛选）")

    pic_root = args.pic_root.resolve()
    total_w = 0
    total_sk = 0
    total_err = 0

    for path in xlsx_files:
        stem = pic_folder_stem_from_path(path)
        sub = mc[mc["pic_folder"].astype(str).str.strip() == stem].copy()
        if sub.empty:
            print(f"[WARN] 映射表无 pic_folder={stem!r}，跳过: {path.name}", file=sys.stderr)
            continue
        sub["_idn"] = pd.to_numeric(sub["id"], errors="coerce").fillna(0).astype(int)
        sub = sub.sort_values("_idn", kind="mergesort").drop(columns=["_idn"])
        sub_use = sub.loc[row_has_new_code(sub["code"])].copy()
        if sub_use.empty:
            print(
                f"[SKIP] {path.name}: 映射中无有效新 code 的行（pic_folder={stem!r}）",
                file=sys.stderr,
            )
            total_err += 1
            continue

        wb = openpyxl.load_workbook(path, data_only=True, rich_text=False)
        ws = wb.active
        code_col = find_code_col(ws)
        if args.match in ("cell", "hybrid") and code_col is None:
            print(f"[ERR] {path.name}: cell 模式需要表头 code 列", file=sys.stderr)
            total_err += 1
            continue

        images = sorted(ws._images, key=image_sort_key)
        n_img, n_row = len(images), len(sub_use)
        if n_img > n_row:
            print(
                f"[WARN] {path.name}: 嵌入图 {n_img} 张 > 有码行 {n_row}，将只导出前 {n_row} 张（请检查是否多图叠在同一行或多余嵌入对象）",
                file=sys.stderr,
            )
        elif n_img < n_row:
            print(
                f"[WARN] {path.name}: 嵌入图 {n_img} 张 < 有码行 {n_row}，将只导出 {n_img} 张",
                file=sys.stderr,
            )
            try:
                disp_n = count_dispimg_formula_cells(path.resolve())
                if disp_n:
                    print(
                        f"[INFO] {path.name}: 表中约 {disp_n} 个单元格含 DISPIMG 公式（WPS 专有），"
                        f"openpyxl 无法当作嵌入图导出，与「仅 {n_img} 张 ws._images」现象一致。",
                        file=sys.stderr,
                    )
            except Exception as e:
                print(f"[INFO] DISPIMG 统计跳过: {e}", file=sys.stderr)
            print(
                "[HINT] 可选做法: (1) 在 WPS/Excel 将图改为真「嵌入单元格」图片后再跑本脚本；"
                "(2) 已有人工对齐的散图则放入 PIC/该分类/，直接命名为 <code>.png/.jpg（与表内 code 一致）；"
                "(3) 或 assign_pics_by_row_order 按 image1… 顺序赋码（易错位，需抽查）。",
                file=sys.stderr,
            )

        dest_dir = pic_root / stem
        if not args.dry_run:
            dest_dir.mkdir(parents=True, exist_ok=True)

        # cell / hybrid：(pic_folder, old_code) -> 按 id 排序的新码队列
        cell_queues: dict[tuple[str, str], deque] = {}
        if args.match in ("cell", "hybrid"):
            for _, r in sub_use.iterrows():
                old_k = str(r["code_before_cleanup"]).strip()
                if not old_k:
                    continue
                key = (stem, old_k)
                cell_queues.setdefault(key, deque()).append(str(r["code"]).strip())

        # hybrid：按 id 顺序尚未分配的新码（fallback 时 popleft）；excel_src_row 行号 -> 新码队列（与锚点行对齐）
        unused_new_codes: deque[str] | None = None
        row_queues: dict[int, deque[str]] = {}
        has_excel_row = False
        if args.match == "hybrid":
            unused_new_codes = deque(str(r["code"]).strip() for _, r in sub_use.iterrows())
            if "excel_src_row" in sub_use.columns:
                ers = sub_use["excel_src_row"].astype(str).str.strip()
                has_excel_row = bool(ers.ne("").any() and ers.str.isdigit().any())
            if has_excel_row:
                sub_sorted = sub_use.assign(
                    _ord=pd.to_numeric(sub_use["id"], errors="coerce").fillna(0).astype(int)
                ).sort_values("_ord", kind="mergesort")
                for _, r in sub_sorted.iterrows():
                    key_r = str(r.get("excel_src_row", "")).strip()
                    if key_r.isdigit():
                        row_queues.setdefault(int(key_r), deque()).append(str(r["code"]).strip())
                print(
                    f"[INFO] {path.name}: hybrid 启用 excel_src_row 锚点行对齐（{len(row_queues)} 个行号有物料）",
                    file=sys.stderr,
                )

        def consume_unused(code: str) -> None:
            if unused_new_codes is None:
                return
            try:
                unused_new_codes.remove(code)
            except ValueError:
                pass

        pos_i = 0
        for i, img in enumerate(images):
            excel_row = int(img.anchor._from.row) + 1
            cell_old = ""
            if code_col is not None:
                cell_code = ws.cell(excel_row, code_col).value
                cell_old = "" if cell_code is None else str(cell_code).strip()

            new_code: str | None = None

            if args.match == "position":
                if pos_i >= len(sub_use):
                    break
                row = sub_use.iloc[pos_i]
                pos_i += 1
                new_code = str(row["code"]).strip()
                old_expect = str(row["code_before_cleanup"]).strip()
                if code_col is not None and old_expect and cell_old and old_expect != cell_old:
                    print(
                        f"[WARN] {path.name} 图#{i + 1} 行{excel_row}: 表内 code={cell_old!r} 与映射 old={old_expect!r} 不一致，仍按位置写入新码 {new_code!r}",
                        file=sys.stderr,
                    )
            elif args.match == "cell":
                key = (stem, cell_old)
                q = cell_queues.get(key)
                if not q:
                    print(
                        f"[SKIP] {path.name} 图#{i + 1} 行{excel_row}: 单元格 code={cell_old!r} 无对应映射队列",
                        file=sys.stderr,
                    )
                    continue
                new_code = q.popleft()
            else:
                # hybrid：1) 锚点行 == excel_src_row 2) 单元格旧码 3) 空单元格按 id 顺序兜底
                rq = row_queues.get(excel_row) if row_queues else None
                if rq:
                    new_code = rq.popleft()
                    consume_unused(new_code)
                else:
                    key = (stem, cell_old)
                    q = cell_queues.get(key)
                    if q:
                        new_code = q.popleft()
                        consume_unused(new_code)
                    elif not cell_old and unused_new_codes is not None and len(unused_new_codes) > 0:
                        new_code = unused_new_codes.popleft()
                        print(
                            f"[INFO] {path.name} 图#{i + 1} 行{excel_row}: 无行号/旧码匹配，hybrid 按 id 序赋新码 {new_code!r}",
                            file=sys.stderr,
                        )
                    else:
                        print(
                            f"[SKIP] {path.name} 图#{i + 1} 行{excel_row}: 单元格 code={cell_old!r} 无对应映射队列且无法 fallback",
                            file=sys.stderr,
                        )
                        continue

            if not new_code:
                continue

            fmt = (getattr(img, "format", None) or "png").lower()
            if fmt == "jpeg":
                ext = ".jpg"
            else:
                ext = "." + fmt if not fmt.startswith(".") else fmt
            if ext not in (".png", ".jpg", ".jpeg", ".webp"):
                ext = ".png"

            out = dest_dir / f"{safe_stem(new_code)}{ext}"
            if args.skip_existing and out.is_file():
                total_sk += 1
                continue
            data = img._data()
            if args.dry_run:
                print(f"  [dry-run] {out}")
                total_w += 1
            else:
                out.write_bytes(data)
                print(f"  [OK] {out.name} <- {path.name} (#{i + 1})")
                total_w += 1

        if args.match in ("cell", "hybrid"):
            bad = {f"{k[1]}": len(v) for k, v in cell_queues.items() if len(v) > 0}
            if bad:
                print(
                    f"[WARN] {path.name}: 下列旧码仍有未分配的新码（图数量不足或锚点行不对）: {bad}",
                    file=sys.stderr,
                )
            if args.match == "hybrid" and unused_new_codes is not None and len(unused_new_codes) > 0:
                print(
                    f"[WARN] {path.name}: hybrid 后仍有 {len(unused_new_codes)} 个新码未对应到图: "
                    f"前若干为 {list(unused_new_codes)[:8]}",
                    file=sys.stderr,
                )

    print(f"\n完成: 写入 {total_w}，跳过已存在 {total_sk}，表级错误 {total_err}")
    if args.dry_run:
        print("(dry-run 未实际写入)")


if __name__ == "__main__":
    main()
