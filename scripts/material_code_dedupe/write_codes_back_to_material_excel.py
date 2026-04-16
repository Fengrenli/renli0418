#!/usr/bin/env python3
"""
将 materials_cleaned.csv 中的新 code 写回 material-excel/*.xlsx（仅改「code」列单元格，尽量保留嵌入图等 OpenXML 结构）。

匹配模式（--match）：
  position  默认。按「非空 name 的数据行顺序」与同 pic_folder 下 CSV 按 id 排序后的行依次对应
            （与 extract_images_from_material_excel.py 的 position 一致；合并后 CSV 的 id 为全局 1..n，
             与各表内 id 往往不一致，故一般不要用 id 模式）。
  id        仅当 Excel 内 id 已与 materials_cleaned 全局 id 一致时使用。
  oldcode   按当前单元格旧 code 匹配 CSV 的 code_before_cleanup，同旧码多行时按 id 升序依次分配新码。

用法：
  python write_codes_back_to_material_excel.py \\
    --excel-dir "C:/Users/Admin/Desktop/material/new/material-excel" \\
    --mapping "E:/renli0418/materials_cleaned.csv" \\
    --match position

  # 合并流水线里的 id 是「全品牌全局」；表内序号列若被误写成全局或跳号，可重排为 1..N：
  python write_codes_back_to_material_excel.py ... --match position --rewrite-local-serial

  # 先备份；试跑不保存：
  python write_codes_back_to_material_excel.py ... --dry-run
"""
from __future__ import annotations

import argparse
import shutil
import sys
from collections import defaultdict, deque
from pathlib import Path
import zipfile

import openpyxl
import pandas as pd


def norm_header(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def col_index_by_header(ws: openpyxl.worksheet.worksheet.Worksheet, want: str) -> int | None:
    for c in range(1, ws.max_column + 1):
        if norm_header(ws.cell(1, c).value) == want:
            return c
    return None


def col_index_any(ws: openpyxl.worksheet.worksheet.Worksheet, wants: tuple[str, ...]) -> int | None:
    want_set = {w.strip().lower() for w in wants if w and w.strip()}
    for c in range(1, ws.max_column + 1):
        h = norm_header(ws.cell(1, c).value)
        if h in want_set:
            return c
    return None


# 表内流水序号（与 merge 时「该 xlsx 内非空 name 行」一致，1..N；勿与合并 CSV 的全局 id 混用）
SERIAL_HEADER_CANDIDATES = ("序号", "编号", "id", "serial", "no")

# 兼容中英表头
CODE_HEADER_CANDIDATES = ("code", "编码", "物料编码", "item code", "itemcode")
NAME_HEADER_CANDIDATES = ("name", "名称", "物料名称", "产品名称", "product name")


def col_index_serial(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    for want in SERIAL_HEADER_CANDIDATES:
        idx = col_index_by_header(ws, want)
        if idx is not None:
            return idx
    return None


def col_index_code(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    return col_index_any(ws, CODE_HEADER_CANDIDATES)


def col_index_name(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    return col_index_any(ws, NAME_HEADER_CANDIDATES)


def pic_folder_stem_from_path(p: Path) -> str:
    s = p.stem.strip().rstrip(".")
    suf = ".bak_before_code"
    if len(s) > len(suf) and s.lower().endswith(suf.lower()):
        return s[: -len(suf)].rstrip(".")
    return s


def is_wps_cellimages_workbook(path: Path) -> bool:
    """若包含 xl/cellimages.xml（WPS 单元格图），openpyxl 保存可能破坏其结构；此类工作簿跳过写回。"""
    try:
        with zipfile.ZipFile(path, "r") as z:
            names = {n.lower() for n in z.namelist()}
            return ("xl/cellimages.xml" in names) or ("xl/cellimages.xml.rels" in names) or ("xl/cellimages.xml.rels".lower() in names)
    except Exception:
        return False


def data_row_indices(ws: openpyxl.worksheet.worksheet.Worksheet, name_col: int) -> list[int]:
    out: list[int] = []
    for r in range(2, ws.max_row + 1):
        nv = ws.cell(r, name_col).value
        if nv is None or str(nv).strip() == "":
            continue
        out.append(r)
    return out


def normalize_id_cell(raw: object) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except ValueError:
        pass
    return s


def main() -> None:
    ap = argparse.ArgumentParser(description="将 materials_cleaned 新码写回 xlsx")
    ap.add_argument(
        "--excel-dir",
        type=Path,
        default=Path(r"C:\Users\Admin\Desktop\material\new\material-excel"),
    )
    ap.add_argument(
        "--mapping",
        type=Path,
        default=Path(__file__).resolve().parents[2] / "materials_cleaned.csv",
        help="materials_cleaned.csv",
    )
    ap.add_argument(
        "--match",
        choices=("id", "position", "oldcode"),
        default="position",
        help="position: 按非空 name 行序对齐；id: 全局 id 一致时用；oldcode: 按单元格旧码对 code_before_cleanup",
    )
    ap.add_argument("--dry-run", action="store_true", help="只打印将要写入的变更，不保存")
    ap.add_argument(
        "--backup-suffix",
        type=str,
        default=".bak_before_code",
        help="非 dry-run 时在同级目录写备份：原名 + 此后缀 + .xlsx",
    )
    ap.add_argument(
        "--rewrite-local-serial",
        action="store_true",
        help="将 序号/id 列按本表非空 name 行重写为 1..N（与 pic_folder 映射行数一致；"
        "仅用于 --match position 或 oldcode）",
    )
    args = ap.parse_args()

    if args.rewrite_local_serial and args.match == "id":
        sys.exit("--rewrite-local-serial 不能与 --match id 同时使用（id 列用于匹配时不能整列重编号）")

    if not args.excel_dir.is_dir():
        sys.exit(f"目录不存在: {args.excel_dir}")
    if not args.mapping.is_file():
        sys.exit(f"找不到: {args.mapping}")

    mc = pd.read_csv(args.mapping, dtype=str, keep_default_na=False)
    mc.columns = [str(c).strip().lower() for c in mc.columns]
    for col in ("id", "code", "pic_folder"):
        if col not in mc.columns:
            sys.exit(f"{args.mapping} 缺少列: {col}")
    if args.match == "oldcode" and "code_before_cleanup" not in mc.columns:
        sys.exit(f"{args.mapping} 缺少 code_before_cleanup 列，不能使用 --match oldcode")

    def sort_sub(df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        out["_idn"] = pd.to_numeric(out["id"], errors="coerce").fillna(0).astype(int)
        return out.sort_values("_idn", kind="mergesort").drop(columns=["_idn"])

    changed = 0
    save_failed: list[str] = []
    files = sorted((p for p in args.excel_dir.glob("*.xlsx") if not p.name.startswith("~$")), key=lambda p: p.name)
    if not files:
        sys.exit(f"未找到 .xlsx: {args.excel_dir}")

    # 同名主表与 bak 同时存在时优先写入 bak（与当前流程一致）
    by_stem: dict[str, list[Path]] = defaultdict(list)
    for p in files:
        by_stem[pic_folder_stem_from_path(p)].append(p)

    chosen_files: list[Path] = []
    for stem_key in sorted(by_stem.keys()):
        grp = sorted(by_stem[stem_key], key=lambda x: x.name.lower())
        baks = [p for p in grp if ".bak_before_code" in p.name.lower()]
        mains = [p for p in grp if ".bak_before_code" not in p.name.lower()]
        chosen_files.append(baks[0] if baks else mains[0])

    for path in chosen_files:
        stem = pic_folder_stem_from_path(path)
        sub = mc[mc["pic_folder"].astype(str).str.strip() == stem].copy()
        if sub.empty:
            print(f"[SKIP] 无 pic_folder={stem!r}: {path.name}")
            continue
        sub = sort_sub(sub)

        if is_wps_cellimages_workbook(path):
            print(f"[SKIP] WPS cellimages 工作簿（为保护 DISPIMG 对齐，不写回）: {path.name}", file=sys.stderr)
            continue

        wb = openpyxl.load_workbook(path, data_only=False, rich_text=False)
        ws = wb.active
        id_col = col_index_serial(ws) or col_index_by_header(ws, "id")
        code_col = col_index_code(ws) or col_index_by_header(ws, "code")
        name_col = col_index_name(ws) or col_index_by_header(ws, "name")
        if code_col is None:
            print(f"[ERR] 无 code 列: {path.name}", file=sys.stderr)
            continue
        if args.match == "id" and id_col is None:
            print(f"[ERR] --match id 需要 id 列: {path.name}", file=sys.stderr)
            continue
        if args.match == "position" and name_col is None:
            print(f"[ERR] --match position 需要 name 列: {path.name}", file=sys.stderr)
            continue
        if args.match == "oldcode" and name_col is None:
            print(f"[ERR] --match oldcode 需要 name 列（用于限定数据行）: {path.name}", file=sys.stderr)
            continue

        id_to_code: dict[str, str] = {}
        for _, r in sub.iterrows():
            k = normalize_id_cell(r["id"])
            if k:
                id_to_code[k] = str(r["code"]).strip()
        codes_in_order = [str(r["code"]).strip() for _, r in sub.iterrows()]

        updates: list[tuple[int, str, str, str]] = []  # excel_row, old, new, how

        if args.match == "id":
            for r in range(2, ws.max_row + 1):
                eid = ws.cell(r, id_col).value
                key = normalize_id_cell(eid)
                if not key:
                    continue
                new_c = id_to_code.get(key)
                if new_c is None:
                    continue
                old_c = ws.cell(r, code_col).value
                old_s = "" if old_c is None else str(old_c).strip()
                if old_s != new_c:
                    updates.append((r, old_s, new_c, f"id={key}"))
        elif args.match == "position":
            excel_rows: list[int] = []
            for r in range(2, ws.max_row + 1):
                nv = ws.cell(r, name_col).value
                if nv is None or str(nv).strip() == "":
                    continue
                excel_rows.append(r)
            if len(excel_rows) != len(codes_in_order):
                print(
                    f"[WARN] {path.name}: 非空 name 行 {len(excel_rows)} vs 映射行 {len(codes_in_order)}，按较短长度对齐",
                    file=sys.stderr,
                )
            for i, r in enumerate(excel_rows):
                if i >= len(codes_in_order):
                    break
                new_c = codes_in_order[i]
                old_c = ws.cell(r, code_col).value
                old_s = "" if old_c is None else str(old_c).strip()
                if old_s != new_c:
                    updates.append((r, old_s, new_c, f"pos#{i + 1}"))
        else:
            # oldcode：单元格当前 code = CSV code_before_cleanup，同旧码多行按 id 顺序出队
            queues: dict[str, deque[str]] = defaultdict(deque)
            for _, r in sub.iterrows():
                o = str(r["code_before_cleanup"]).strip()
                if not o:
                    continue
                queues[o].append(str(r["code"]).strip())
            for r in range(2, ws.max_row + 1):
                nv = ws.cell(r, name_col).value
                if nv is None or str(nv).strip() == "":
                    continue
                old_c = ws.cell(r, code_col).value
                old_s = "" if old_c is None else str(old_c).strip()
                if not old_s:
                    continue
                q = queues.get(old_s)
                if not q:
                    print(
                        f"[WARN] {path.name} 行{r}: 单元格 code={old_s!r} 在映射中无 code_before_cleanup 队列（可能已是新码）",
                        file=sys.stderr,
                    )
                    continue
                new_c = q.popleft()
                if old_s != new_c:
                    updates.append((r, old_s, new_c, "oldcode"))
            leftover = {k: len(v) for k, v in queues.items() if len(v) > 0}
            if leftover:
                print(
                    f"[WARN] {path.name}: 下列旧码仍有未分配的新码（行数不足或表内已改码）: {leftover}",
                    file=sys.stderr,
                )

        serial_updates: list[tuple[int, object, int]] = []
        serial_col_resolved: int | None = None
        if args.rewrite_local_serial and args.match in ("position", "oldcode"):
            serial_col_resolved = col_index_serial(ws)
            if serial_col_resolved is None:
                print(
                    f"[WARN] {path.name}: --rewrite-local-serial 但未找到列头 序号/编号/id/serial/no，跳过重排",
                    file=sys.stderr,
                )
            else:
                excel_rows = data_row_indices(ws, name_col)
                if args.match == "position" and len(excel_rows) != len(codes_in_order):
                    print(
                        f"[WARN] {path.name}: 非空 name 行 {len(excel_rows)} 与映射 {len(codes_in_order)} 行不齐，"
                        f"序号仍按表内 {len(excel_rows)} 行写 1..{len(excel_rows)}",
                        file=sys.stderr,
                    )
                for i, r in enumerate(excel_rows, start=1):
                    old_v = ws.cell(r, serial_col_resolved).value
                    if old_v is None or str(old_v).strip() == "":
                        old_disp: object = old_v
                    else:
                        try:
                            old_disp = int(float(str(old_v).strip()))
                        except ValueError:
                            old_disp = old_v
                    if old_disp != i:
                        serial_updates.append((r, old_disp, i))

        if not updates and not serial_updates:
            print(f"[OK] 无需修改: {path.name}")
            continue

        for r, old_s, new_c, how in updates:
            print(f"  {path.name} 行{r} [{how}] {old_s!r} -> {new_c!r}")
            changed += 1
            if not args.dry_run:
                ws.cell(r, code_col, value=new_c)

        for r, old_v, new_i in serial_updates:
            print(f"  {path.name} 行{r} [序号] {old_v!r} -> {new_i}")
            changed += 1
            if not args.dry_run and serial_col_resolved is not None:
                ws.cell(r, serial_col_resolved, value=new_i)

        if args.dry_run:
            print(f"  (dry-run 未保存 {path.name})")
        else:
            backup = path.with_name(path.stem + args.backup_suffix + path.suffix)
            try:
                shutil.copy2(path, backup)
                wb.save(path)
                print(f"  已保存 {path.name}，备份: {backup.name}")
            except PermissionError:
                save_failed.append(path.name)
                print(
                    f"[ERR] 保存失败（文件可能正在被 WPS/Excel 占用）: {path.name}。请关闭该文件后重跑。",
                    file=sys.stderr,
                )
            except Exception as e:
                save_failed.append(path.name)
                print(f"[ERR] 保存失败 {path.name}: {e}", file=sys.stderr)

    print(f"\n合计将更新/已更新单元格: {changed}")
    if save_failed:
        print(f"[WARN] 以下文件未保存成功（本次已跳过，不影响其它文件）: {', '.join(save_failed)}", file=sys.stderr)
        raise SystemExit("存在未能保存的工作簿（多为被 WPS/Excel 占用）。请关闭这些文件后重跑。")


if __name__ == "__main__":
    main()
